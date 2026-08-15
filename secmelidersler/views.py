from django.contrib import messages
from django.db.models import Count, Max, Value
from django.urls import reverse
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from okul.auth import mudur_yardimcisi_required

from .forms import AlanForm, OrtakDersHavuzuForm, SecmeliDersForm, SecmeliDersGrubuForm, SecmeliDersHavuzuForm, SinifSeviyeToplamSaatForm
from .models import (
    Alan, AlanDers, OgrenciOrtalama, OgrenciSinifTekrari, OgrenciTasdikname,
    OrtakDers, OrtakDersHavuzu,
    SecmeliDers, SecmeliDersGrubu, SecmeliDersHavuzu,
    SinifSeviyeToplamSaat,
    get_aktif_egitim_yili, get_toplam_saat,
)
from okul.models import EgitimOgretimYili
from .services import donem_kopyala
from .services.ders_dagilimi import baskin_egitim_yili, plan_sinif_dagilimi, plan_sinif_dagilimi_gecmis

_SINIFLAR = [9, 10, 11, 12]


def _yf(qs, aktif_yil):
    """Aktif EÖY varsa queryset'i yıla göre filtreler; yoksa tüm kayıtları döndürür."""
    return qs.filter(egitim_yili=aktif_yil) if aktif_yil else qs


def _secili_yil(request, aktif_yil):
    """?yil=<pk> (GET'te ya da POST body'sindeki gizli alanda) verilmişse o
    EgitimOgretimYili'ni, yoksa aktif_yil'i döner."""
    yil_pk = (request.GET.get("yil") or request.POST.get("yil") or "").strip()
    if yil_pk:
        secili = EgitimOgretimYili.objects.filter(pk=yil_pk).first()
        if secili:
            return secili
    return aktif_yil


@mudur_yardimcisi_required
def index(request):
    aktif_yil = get_aktif_egitim_yili()
    return render(request, "secmelidersler/index.html", {
        "title": "Seçmeli Dersler Yönetimi",
        "aktif_yil": aktif_yil,
    })


@mudur_yardimcisi_required
def alan_listesi(request):
    aktif_yil = get_aktif_egitim_yili()
    secili_yil = _secili_yil(request, aktif_yil)
    alanlar = _yf(Alan.objects.all(), secili_yil).prefetch_related("dersler").order_by("sinif_seviyesi", "sira")
    alan_map = {}
    for a in alanlar:
        alan_map.setdefault(a.sinif_seviyesi, []).append(a)
    sinif_alanlari = [(sv, alan_map.get(sv, [])) for sv in _SINIFLAR]
    return render(request, "secmelidersler/alan_listesi.html", {
        "title": "Alan Tanımları",
        "sinif_alanlari": sinif_alanlari,
        "alan_var": bool(alan_map),
        "aktif_yil": secili_yil,
        "tum_yillar": EgitimOgretimYili.objects.all(),
        "secili_yil": secili_yil,
        "salt_okunur": secili_yil != aktif_yil,
    })


@mudur_yardimcisi_required
def alan_form(request, pk=None):
    aktif_yil = get_aktif_egitim_yili()
    if pk:
        alan = get_object_or_404(Alan, pk=pk)
        title = f"Alan Düzenle — {alan.sinif_seviyesi}. Sınıf / {alan.adi}"
        # Düzenlenen alanın kendi yılı esas alınır — hangi ?yil= ile buraya gelindiğinden
        # bağımsız olarak doğru kataloğu (ders seçenekleri) gösterir.
        secili_yil = alan.egitim_yili or _secili_yil(request, aktif_yil)
    else:
        alan = None
        title = "Yeni Alan Tanımla"
        # Yeni alan, hangi yıl görüntülenirken oluşturulduysa o yıla (secili_yil) etiketlenir —
        # bkz. secmeli_grup_form'daki aynı gerekçe.
        secili_yil = _secili_yil(request, aktif_yil)

    sinif_ozet = {}
    for sv in (9, 10, 11, 12):
        toplam = get_toplam_saat(sv, egitim_yili=secili_yil)
        ortak_dersler = list(
            _yf(OrtakDers.objects.filter(sinif_seviyesi=sv), secili_yil).order_by("sira")
        )
        ortak = sum(od.haftalik_saat for od in ortak_dersler)
        sinif_ozet[sv] = {
            "toplam": toplam,
            "ortak": ortak,
            "ortak_dersler": ortak_dersler,
            "secmeli_maks": toplam - ortak,
        }

    if request.method == "POST":
        form = AlanForm(request.POST, instance=alan, egitim_yili=secili_yil)
        if form.is_valid():
            if not form.instance.pk:
                form.instance.egitim_yili = secili_yil
            form.save()
            messages.success(request, "Alan kaydedildi.")
            yonlendir = reverse("secmeli_alan_listesi")
            if secili_yil:
                yonlendir += f"?yil={secili_yil.pk}"
            return redirect(yonlendir)
    else:
        form = AlanForm(instance=alan, egitim_yili=secili_yil)

    return render(request, "secmelidersler/alan_form.html", {
        "title": title,
        "form": form,
        "alan": alan,
        "sinif_ozet": sinif_ozet,
        "aktif_yil": aktif_yil,
        "secili_yil": secili_yil,
    })


@mudur_yardimcisi_required
def alan_sil(request, pk):
    alan = get_object_or_404(Alan, pk=pk)
    if request.method == "POST":
        alan.delete()
        messages.success(request, f"'{alan.adi}' alanı silindi.")
    yil_param = request.POST.get("yil") or request.GET.get("yil", "")
    yonlendir = reverse("secmeli_alan_listesi")
    if yil_param:
        yonlendir += f"?yil={yil_param}"
    return redirect(yonlendir)


@mudur_yardimcisi_required
def secmeli_grup_listesi(request):
    aktif_yil = get_aktif_egitim_yili()
    secili_yil = _secili_yil(request, aktif_yil)
    gruplar = _yf(
        SecmeliDersGrubu.objects.annotate(ders_sayisi=Count("dersler")),
        secili_yil,
    ).order_by("sinif_seviyesi", "sira")

    sinif_map = {}
    for g in gruplar:
        sinif_map.setdefault(g.sinif_seviyesi, []).append(g)
    sinif_gruplari = [(s, sinif_map.get(s, [])) for s in _SINIFLAR]
    return render(request, "secmelidersler/secmeli_grup_listesi.html", {
        "title": "Seçmeli Ders Grupları",
        "sinif_gruplari": sinif_gruplari,
        "aktif_yil": secili_yil,
        "tum_yillar": EgitimOgretimYili.objects.all(),
        "secili_yil": secili_yil,
        "salt_okunur": secili_yil != aktif_yil,
    })


@mudur_yardimcisi_required
def secmeli_grup_form(request, pk=None):
    aktif_yil = get_aktif_egitim_yili()
    # Yeni grup, hangi yıl görüntülenirken oluşturulduysa o yıla (secili_yil) etiketlenir —
    # her zaman aktif_yil'e sabitlersek geçmiş yıl (?yil=) görüntülenirken oluşturulan bir
    # grup yanlış yıla bağlanır (bkz. secmeli_grup_listesi — artık yıl kilidi yok, her iki
    # yılda da grup/ders oluşturulabilir).
    secili_yil = _secili_yil(request, aktif_yil)
    if pk:
        grup = get_object_or_404(SecmeliDersGrubu, pk=pk)
        title = f"Grup Düzenle — {grup.sinif_seviyesi}. Sınıf / {grup.adi}"
    else:
        grup = None
        title = "Yeni Seçmeli Ders Grubu"

    if request.method == "POST":
        form = SecmeliDersGrubuForm(request.POST, instance=grup)
        if form.is_valid():
            yeni_grup = form.save(commit=False)
            if not yeni_grup.pk:
                yeni_grup.egitim_yili = secili_yil
            yeni_grup.save()
            messages.success(request, "Grup kaydedildi.")
            yonlendir = reverse("secmeli_grup_listesi")
            if secili_yil:
                yonlendir += f"?yil={secili_yil.pk}"
            return redirect(yonlendir)
    else:
        initial = {}
        sinif_param = request.GET.get("sinif", "").strip()
        if sinif_param.isdigit() and int(sinif_param) in _SINIFLAR:
            initial["sinif_seviyesi"] = int(sinif_param)
        form = SecmeliDersGrubuForm(instance=grup, initial=initial or None)

    return render(request, "secmelidersler/secmeli_grup_form.html", {
        "title": title,
        "form": form,
        "grup": grup,
        "aktif_yil": aktif_yil,
        "secili_yil": secili_yil,
    })


@mudur_yardimcisi_required
def secmeli_grup_sil(request, pk):
    grup = get_object_or_404(SecmeliDersGrubu, pk=pk)
    if request.method == "POST":
        grup.delete()
        messages.success(request, f"'{grup.adi}' grubu ve tüm dersleri silindi.")
    yil_param = request.POST.get("yil") or request.GET.get("yil", "")
    yonlendir = reverse("secmeli_grup_listesi")
    if yil_param:
        yonlendir += f"?yil={yil_param}"
    return redirect(yonlendir)


@mudur_yardimcisi_required
def secmeli_ders_havuzdan_ekle(request, grup_pk):
    if request.method != "POST":
        return redirect("secmeli_grup_ders_listesi", grup_pk=grup_pk)

    grup = get_object_or_404(SecmeliDersGrubu, pk=grup_pk)
    ders_adi = request.POST.get("ders_adi", "").strip()
    saat_secenekleri = request.POST.get("saat_secenekleri", "").strip()

    if not ders_adi or not saat_secenekleri:
        messages.error(request, "Geçersiz istek.")
        return redirect("secmeli_grup_ders_listesi", grup_pk=grup_pk)

    if SecmeliDers.objects.filter(grup=grup, ders_adi=ders_adi).exists():
        messages.warning(request, f"'{ders_adi}' bu grupta zaten mevcut.")
    else:
        maks_sira = SecmeliDers.objects.filter(grup=grup).aggregate(
            m=Coalesce(Max("sira"), Value(0))
        )["m"]
        SecmeliDers.objects.create(
            grup=grup,
            ders_adi=ders_adi,
            saat_secenekleri=saat_secenekleri,
            sira=maks_sira + 1,
            aktif=True,
        )
        messages.success(request, f"'{ders_adi}' gruba eklendi.")

    return redirect("secmeli_grup_ders_listesi", grup_pk=grup_pk)


@mudur_yardimcisi_required
def secmeli_ders_form(request, grup_pk, pk=None):
    grup = get_object_or_404(SecmeliDersGrubu, pk=grup_pk)
    if pk:
        ders = get_object_or_404(SecmeliDers, pk=pk, grup=grup)
        title = f"Ders Düzenle — {ders.ders_adi}"
    else:
        ders = None
        title = f"Yeni Ders — {grup.sinif_seviyesi}. Sınıf / {grup.adi}"

    if request.method == "POST":
        form = SecmeliDersForm(request.POST, instance=ders)
        if form.is_valid():
            yeni_ders = form.save(commit=False)
            yeni_ders.grup = grup
            yeni_ders.save()
            messages.success(request, "Ders kaydedildi.")
            return redirect("secmeli_grup_ders_listesi", grup_pk=grup_pk)
    else:
        initial = {}
        ders_adi_param = request.GET.get("ders_adi", "").strip()
        if ders_adi_param:
            initial["ders_adi"] = ders_adi_param
        form = SecmeliDersForm(instance=ders, initial=initial or None)

    return render(request, "secmelidersler/secmeli_ders_form.html", {
        "title": title,
        "form": form,
        "grup": grup,
        "ders": ders,
        "saat_secenek_araligi": range(1, 13),
    })


@mudur_yardimcisi_required
def secmeli_ders_sil(request, grup_pk, pk):
    grup = get_object_or_404(SecmeliDersGrubu, pk=grup_pk)
    ders = get_object_or_404(SecmeliDers, pk=pk, grup=grup)
    if request.method == "POST":
        ders.delete()
        messages.success(request, f"'{ders.ders_adi}' dersi silindi.")
    return redirect("secmeli_grup_ders_listesi", grup_pk=grup_pk)


@mudur_yardimcisi_required
def secmeli_grup_ders_listesi(request, grup_pk):
    grup = get_object_or_404(SecmeliDersGrubu, pk=grup_pk)

    grup_dersleri = list(grup.dersler.order_by("sira"))
    grup_adlari = {d.ders_adi for d in grup_dersleri}

    havuz_dersler = list(SecmeliDersHavuzu.objects.all())
    havuz_adlari = {h.ders_adi for h in havuz_dersler}
    havuz_map = {h.ders_adi: h for h in havuz_dersler}

    sag_kart = []

    # 1. Grupta olan dersler — atanmış olarak göster (Çıkar)
    for d in grup_dersleri:
        sag_kart.append({
            "ders_adi": d.ders_adi,
            "atanmis": True,
            "aktif": d.aktif,
            "ders_obj": d,
            "havuz_ders": havuz_map.get(d.ders_adi),
            "saat_str": d.saat_secenekleri,
        })

    # 2. Havuzda olup grupta olmayan dersler — Ekle
    for h in havuz_dersler:
        if h.ders_adi not in grup_adlari:
            sag_kart.append({
                "ders_adi": h.ders_adi,
                "atanmis": False,
                "aktif": h.aktif,
                "ders_obj": None,
                "havuz_ders": h,
                "saat_str": h.derssaati,
            })

    return render(request, "secmelidersler/secmeli_grup_ders_listesi.html", {
        "title": f"{grup.sinif_seviyesi}. Sınıf — {grup.adi}",
        "grup": grup,
        "grup_dersleri": grup_dersleri,
        "havuz": sag_kart,
        "veri_var": bool(grup_dersleri),
    })


@mudur_yardimcisi_required
def ortak_ders_listesi(request):
    aktif_yil = get_aktif_egitim_yili()
    secili_yil = _secili_yil(request, aktif_yil)
    salt_okunur = secili_yil != aktif_yil
    tum_yillar = EgitimOgretimYili.objects.all()
    yil_qs = f"&yil={secili_yil.pk}" if salt_okunur and secili_yil else ""
    sinif_param = request.GET.get("sinif", "").strip()
    secili_sinif = int(sinif_param) if sinif_param.isdigit() and int(sinif_param) in _SINIFLAR else None

    ders_qs = _yf(OrtakDers.objects.order_by("sinif_seviyesi", "sira"), secili_yil)
    ders_map = {}
    ders_sira = {}
    for ders in ders_qs:
        if ders.ders_adi not in ders_map:
            ders_map[ders.ders_adi] = {}
            ders_sira[ders.ders_adi] = ders.sira
        ders_map[ders.ders_adi][ders.sinif_seviyesi] = ders
        ders_sira[ders.ders_adi] = min(ders_sira[ders.ders_adi], ders.sira)

    ders_adlari = sorted(ders_map, key=lambda d: ders_sira[d])

    if secili_sinif:
        sinif_dersleri = [
            ders_map[adi][secili_sinif]
            for adi in ders_adlari
            if secili_sinif in ders_map[adi]
        ]
        sinif_ders_adlari = {d.ders_adi for d in sinif_dersleri}
        havuz = [
            {
                "havuz_ders": hd,
                "ders_adi": hd.ders_adi,
                "aktif": hd.aktif,
                "atanmis": hd.ders_adi in sinif_ders_adlari,
                "ders_obj": ders_map.get(hd.ders_adi, {}).get(secili_sinif),
            }
            for hd in OrtakDersHavuzu.objects.all()
        ]
        return render(request, "secmelidersler/ortak_ders_listesi.html", {
            "title": f"Zorunlu Dersler — {secili_sinif}. Sınıf",
            "siniflar": _SINIFLAR,
            "secili_sinif": secili_sinif,
            "sinif_dersleri": sinif_dersleri,
            "sinif_toplam": sum(d.haftalik_saat for d in sinif_dersleri),
            "havuz": havuz,
            "veri_var": bool(sinif_dersleri),
            "aktif_yil": secili_yil,
            "tum_yillar": tum_yillar,
            "secili_yil": secili_yil,
            "salt_okunur": salt_okunur,
            "yil_qs": yil_qs,
        })

    satirlar = [(adi, [ders_map[adi].get(s) for s in _SINIFLAR]) for adi in ders_adlari]
    toplamlar = [
        sum(ders_map[adi].get(s).haftalik_saat for adi in ders_map if s in ders_map[adi])
        for s in _SINIFLAR
    ]
    return render(request, "secmelidersler/ortak_ders_listesi.html", {
        "title": "Zorunlu (Ortak) Dersler",
        "siniflar": _SINIFLAR,
        "secili_sinif": None,
        "satirlar": satirlar,
        "toplamlar": toplamlar,
        "veri_var": bool(satirlar),
        "aktif_yil": secili_yil,
        "tum_yillar": tum_yillar,
        "secili_yil": secili_yil,
        "salt_okunur": salt_okunur,
        "yil_qs": yil_qs,
    })


@mudur_yardimcisi_required
def ortak_ders_sil(request, pk):
    ders = get_object_or_404(OrtakDers, pk=pk)
    if request.method == "POST":
        ders.delete()
        messages.success(request, f"'{ders.ders_adi}' dersi silindi.")
        ref_sinif = request.POST.get("ref_sinif", "").strip()
        if ref_sinif.isdigit() and int(ref_sinif) in _SINIFLAR:
            return redirect(f"{reverse('secmeli_ortak_ders_listesi')}?sinif={ref_sinif}")
    return redirect("secmeli_ortak_ders_listesi")


@mudur_yardimcisi_required
def ortak_ders_havuzdan_ekle(request):
    if request.method != "POST":
        return redirect("secmeli_ortak_ders_listesi")

    aktif_yil = get_aktif_egitim_yili()
    sinif_str = request.POST.get("sinif_seviyesi", "").strip()
    ders_adi = request.POST.get("ders_adi", "").strip()
    saat_str = request.POST.get("haftalik_saat", "").strip()

    secili_sinif = int(sinif_str) if sinif_str.isdigit() and int(sinif_str) in _SINIFLAR else None
    if not secili_sinif or not ders_adi:
        messages.error(request, "Geçersiz istek.")
        return redirect("secmeli_ortak_ders_listesi")

    try:
        haftalik_saat = int(saat_str.split(",")[0].strip())
        if haftalik_saat < 1:
            raise ValueError
    except (ValueError, TypeError):
        messages.error(request, "Geçerli bir haftalık saat girin (en az 1).")
        return redirect(f"{reverse('secmeli_ortak_ders_listesi')}?sinif={secili_sinif}")

    kontrol_qs = _yf(
        OrtakDers.objects.filter(sinif_seviyesi=secili_sinif, ders_adi=ders_adi),
        aktif_yil,
    )
    if kontrol_qs.exists():
        messages.warning(request, f"'{ders_adi}' bu sınıf seviyesi ve yıl için zaten tanımlı.")
    else:
        maks_sira = _yf(OrtakDers.objects.filter(sinif_seviyesi=secili_sinif), aktif_yil).aggregate(
            m=Coalesce(Max("sira"), Value(0))
        )["m"]
        OrtakDers.objects.create(
            egitim_yili=aktif_yil,
            sinif_seviyesi=secili_sinif,
            ders_adi=ders_adi,
            haftalik_saat=haftalik_saat,
            sira=maks_sira + 1,
        )
        messages.success(request, f"'{ders_adi}' {secili_sinif}. sınıfa eklendi.")

    return redirect(f"{reverse('secmeli_ortak_ders_listesi')}?sinif={secili_sinif}")


# ── Seçmeli Ders Havuzu CRUD ────────────────────────────────────────────────

@mudur_yardimcisi_required
def secmeli_havuz_listesi(request):
    dersler = SecmeliDersHavuzu.objects.prefetch_related("branslar")
    return render(request, "secmelidersler/secmeli_havuz_listesi.html", {
        "title": "Seçmeli Ders Havuzu",
        "dersler": dersler,
    })


@mudur_yardimcisi_required
def secmeli_havuz_form(request, pk=None):
    if pk:
        ders = get_object_or_404(SecmeliDersHavuzu, pk=pk)
        title = f"Havuz Dersi Düzenle — {ders.ders_adi}"
    else:
        ders = None
        title = "Havuza Yeni Ders Ekle"

    if request.method == "POST":
        form = SecmeliDersHavuzuForm(request.POST, instance=ders)
        if form.is_valid():
            form.save()
            messages.success(request, "Ders kaydedildi.")
            return redirect("secmeli_havuz_listesi")
    else:
        form = SecmeliDersHavuzuForm(instance=ders)

    return render(request, "secmelidersler/secmeli_havuz_form.html", {
        "title": title,
        "form": form,
        "ders": ders,
    })


@mudur_yardimcisi_required
def secmeli_havuz_sil(request, pk):
    ders = get_object_or_404(SecmeliDersHavuzu, pk=pk)
    if request.method == "POST":
        ders.delete()
        messages.success(request, f"'{ders.ders_adi}' havuzdan silindi.")
    return redirect("secmeli_havuz_listesi")


# ── Ortak Ders Havuzu CRUD ──────────────────────────────────────────────────

@mudur_yardimcisi_required
def ortak_havuz_listesi(request):
    dersler = OrtakDersHavuzu.objects.prefetch_related("branslar")
    return render(request, "secmelidersler/ortak_havuz_listesi.html", {
        "title": "Ortak (Zorunlu) Ders Havuzu",
        "dersler": dersler,
    })


@mudur_yardimcisi_required
def ortak_havuz_form(request, pk=None):
    if pk:
        ders = get_object_or_404(OrtakDersHavuzu, pk=pk)
        title = f"Havuz Dersi Düzenle — {ders.ders_adi}"
    else:
        ders = None
        title = "Havuza Yeni Ders Ekle"

    if request.method == "POST":
        form = OrtakDersHavuzuForm(request.POST, instance=ders)
        if form.is_valid():
            form.save()
            messages.success(request, "Ders kaydedildi.")
            return redirect("ortak_havuz_listesi")
    else:
        form = OrtakDersHavuzuForm(instance=ders)

    return render(request, "secmelidersler/ortak_havuz_form.html", {
        "title": title,
        "form": form,
        "ders": ders,
    })


@mudur_yardimcisi_required
def ortak_havuz_sil(request, pk):
    ders = get_object_or_404(OrtakDersHavuzu, pk=pk)
    if request.method == "POST":
        ders.delete()
        messages.success(request, f"'{ders.ders_adi}' havuzdan silindi.")
    return redirect("ortak_havuz_listesi")


# ── Sınıf Seviyesi Toplam Saat ──────────────────────────────────────────────

@mudur_yardimcisi_required
def sinif_toplam_saat_listesi(request):
    aktif_yil = get_aktif_egitim_yili()

    if request.method == "POST":
        pk = request.POST.get("pk")
        obj = get_object_or_404(SinifSeviyeToplamSaat, pk=pk)
        form = SinifSeviyeToplamSaatForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f"{obj.sinif_seviyesi}. sınıf toplam saati güncellendi.")
        else:
            for hata in form.errors.values():
                messages.error(request, hata.as_text())
        return redirect("secmeli_sinif_toplam_saat")

    mevcut = {
        obj.sinif_seviyesi: obj
        for obj in _yf(SinifSeviyeToplamSaat.objects.all(), aktif_yil)
    }
    satirlar = []
    for sv in _SINIFLAR:
        if sv not in mevcut:
            obj = SinifSeviyeToplamSaat.objects.create(
                egitim_yili=aktif_yil,
                sinif_seviyesi=sv,
                haftalik_toplam_saat=40,
            )
            mevcut[sv] = obj
        obj = mevcut[sv]
        satirlar.append({
            "obj": obj,
            "form": SinifSeviyeToplamSaatForm(instance=obj),
        })

    return render(request, "secmelidersler/sinif_toplam_saat_listesi.html", {
        "title": "Sınıf Seviyesi Toplam Ders Saati",
        "satırlar": satirlar,
        "aktif_yil": aktif_yil,
    })


@mudur_yardimcisi_required
def sinif_dagilimi(request):
    aktif_yil = get_aktif_egitim_yili()
    secili_yil = _secili_yil(request, aktif_yil)
    # Bir yıl açıkça seçilmişse (aktif yıl dahil) denetim modu: şu an fiilen o seviyede
    # olan öğrencileri seçilen yılın Alan kayıtlarıyla eşleştirir. Hiç seçim yapılmamışsa
    # (varsayılan sayfa açılışı) eski ileriye dönük "plan" görünümü kullanılır.
    denetim_modu = bool(request.GET.get("yil", "").strip())
    gecmis_yil = secili_yil != aktif_yil

    if denetim_modu:
        # 9. ve 10. sınıfta gerçek bir Alan (izlence) ayrımı yok — yalnızca "YOK" adlı
        # tek bir yer tutucu Alan/AlanDers kaydı var. Yine de denetim ekranının geri
        # kalanıyla (11./12. sınıf) aynı Alan-eşleştirme mekanizması burada da
        # kullanılır; ileriye dönük ("plan") modda bu iki seviye için doğal bir
        # kaynak kohort (8. sınıftan gelen ya da 9. sınıfa alan seçimiyle giren
        # öğrenci) olmadığından yalnızca denetim modunda gösterilir.
        plan_9 = plan_sinif_dagilimi_gecmis(9, secili_yil)
        plan_10 = plan_sinif_dagilimi_gecmis(10, secili_yil)
        plan_11 = plan_sinif_dagilimi_gecmis(11, secili_yil)
        plan_12 = plan_sinif_dagilimi_gecmis(12, secili_yil)
    else:
        plan_9 = None
        plan_10 = None
        plan_11 = plan_sinif_dagilimi(10, 11, aktif_yil)
        plan_12 = plan_sinif_dagilimi(11, 12, aktif_yil)

    # Alan-değiştir butonları, hangi yılın denetim kartı görüntülenirse görüntülensin,
    # o sınıf seviyesindeki kohortun ÇOĞUNLUĞUNUN zaten bağlı olduğu yılın Alan kayıtlarını
    # kullanır (kayıt yoksa aktif yıla düşer) — bkz. `baskin_egitim_yili`. Bu, YUKARIDAKİ
    # `plan_sinif_dagilimi_gecmis` eşleştirmesiyle AYNI kaynağı kullanmalı; aksi hâlde
    # (örn. hep aktif yıl kullansaydık) 11. sınıf gibi seçimi bir önceki yıl yapılmış bir
    # kohortta yapılacak bir atama, o öğrenciyi kohortun geri kalanından farklı bir yılın
    # kataloğuna bağlar — tutarsızlık yaratır (yaşandı: bkz. commit geçmişi).
    aktif_alanlar_9 = list(
        Alan.objects.filter(sinif_seviyesi=9, egitim_yili=baskin_egitim_yili(9, aktif_yil)).order_by("sira")
    )
    aktif_alanlar_10 = list(
        Alan.objects.filter(sinif_seviyesi=10, egitim_yili=baskin_egitim_yili(10, aktif_yil)).order_by("sira")
    )
    aktif_alanlar_11 = list(
        Alan.objects.filter(sinif_seviyesi=11, egitim_yili=baskin_egitim_yili(11, aktif_yil)).order_by("sira")
    )
    aktif_alanlar_12 = list(
        Alan.objects.filter(sinif_seviyesi=12, egitim_yili=baskin_egitim_yili(12, aktif_yil)).order_by("sira")
    )

    return render(request, "secmelidersler/sinif_dagilimi.html", {
        "title": "Sınıf Dağılımı",
        "aktif_yil": aktif_yil,
        "plan_9": plan_9,
        "plan_10": plan_10,
        "plan_11": plan_11,
        "plan_12": plan_12,
        "tum_yillar": EgitimOgretimYili.objects.all(),
        "secili_yil": secili_yil,
        "salt_okunur": denetim_modu,
        "gecmis_yil": gecmis_yil,
        "aktif_alanlar_9": aktif_alanlar_9,
        "aktif_alanlar_10": aktif_alanlar_10,
        "aktif_alanlar_11": aktif_alanlar_11,
        "aktif_alanlar_12": aktif_alanlar_12,
    })


@mudur_yardimcisi_required
@require_POST
def sinif_dagilimi_alan_degistir(request, ogrenci_pk, alan_pk):
    """Sınıf Dağılımı (geçmiş yıl denetimi) ekranından bir öğrencinin seçmeli ders
    seçimini, seçilen Alan'ın (o yıla ait olabilecek) ders paketiyle değiştirir.
    `ogrencidersleri.views.ogrenci_alan_ata` ile aynı mantığı kullanır."""
    from ogrenci.models import Ogrenci
    from ogrencidersleri.models import OgrenciSecmeliDers

    ogrenci = get_object_or_404(Ogrenci, pk=ogrenci_pk)
    alan = get_object_or_404(Alan, pk=alan_pk)

    alan_dersler = list(AlanDers.objects.filter(alan=alan).select_related("ders"))
    if not alan_dersler:
        messages.error(request, f"{alan.adi} alanı için tanımlı ders bulunmuyor.")
    else:
        from .services.secim_sayisi import secim_sayisi_asim_uyarilari
        secimler = [(ad.ders, ad.secilen_saat) for ad in alan_dersler]
        uyarilar = secim_sayisi_asim_uyarilari(ogrenci, secimler, haric_egitim_yili=alan.egitim_yili)

        OgrenciSecmeliDers.objects.filter(
            ogrenci=ogrenci,
            ders__grup__sinif_seviyesi=alan.sinif_seviyesi,
            ders__grup__egitim_yili=alan.egitim_yili,
        ).delete()
        OgrenciSecmeliDers.objects.bulk_create([
            OgrenciSecmeliDers(ogrenci=ogrenci, ders=ad.ders, secilen_saat=ad.secilen_saat)
            for ad in alan_dersler
        ])
        if ogrenci.sinif == alan.sinif_seviyesi:
            ogrenci.sectigi_alan = alan.adi
            ogrenci.save(update_fields=["sectigi_alan"])
        messages.success(
            request,
            f"{ogrenci.adi} {ogrenci.soyadi} — seçmeli ders seçimi {alan.adi} alanına göre güncellendi.",
        )
        for uyari in uyarilar:
            messages.warning(request, uyari)

    yil_pk = request.POST.get("yil", "").strip()
    url = reverse("secmeli_sinif_dagilimi")
    if yil_pk:
        url += f"?yil={yil_pk}"
    return redirect(url)


# ---------------------------------------------------------------------------
# Öğrenci Dönem Ağırlıklı Ortalaması — CRUD
# ---------------------------------------------------------------------------

def _excel_satirlari_oku(dosya):
    """
    XLS veya XLSX dosyasından tüm satırları liste olarak döndürür.
    Her satır: [hücre1, hücre2, ...] (ham değer).
    İlk satır başlık satırı olarak döndürülür.
    """
    adi = dosya.name.lower()
    if adi.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(dosya, read_only=True, data_only=True)
        ws = wb.active
        satirlar = [[c.value for c in row] for row in ws.iter_rows()]
        wb.close()
    elif adi.endswith(".xls"):
        import xlrd, tempfile, os
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
        for chunk in dosya.chunks():
            tmp.write(chunk)
        tmp.close()
        wb = xlrd.open_workbook(tmp.name)
        ws = wb.sheet_by_index(0)
        satirlar = [ws.row_values(r) for r in range(ws.nrows)]
        os.unlink(tmp.name)
    else:
        raise ValueError("Desteklenmeyen dosya formatı. Lütfen .xls veya .xlsx yükleyin.")
    return satirlar


def _sutun_indeksleri_bul(baslik_satiri):
    """Başlık satırından okulno ve a_ortalama sütun indekslerini bulur."""
    normalize = lambda s: str(s).strip().lower().replace(" ", "_").replace("ı", "i").replace("ö", "o").replace("ü", "u")
    basliklar = [normalize(h) for h in baslik_satiri]
    okulno_idx = a_ort_idx = None
    for i, h in enumerate(basliklar):
        if "okulno" in h or h == "okul_no":
            okulno_idx = i
        if "a_ortalama" in h or "agirlikli" in h or "ortalama" in h:
            a_ort_idx = i
    return okulno_idx, a_ort_idx


@mudur_yardimcisi_required
def ortalama_listesi(request):
    from django.core.paginator import Paginator

    aktif_yil = get_aktif_egitim_yili()
    qs = (
        OgrenciOrtalama.objects
        .filter(egitim_yili=aktif_yil)
        .select_related("ogrenci")
        .order_by("ogrenci__okulno")
    )

    arama = request.GET.get("q", "").strip()
    if arama:
        from django.db.models import Q
        qs = qs.filter(
            Q(ogrenci__okulno__icontains=arama)
            | Q(ogrenci__adi__icontains=arama)
            | Q(ogrenci__soyadi__icontains=arama)
        )

    paginator = Paginator(qs, 50)
    sayfa = paginator.get_page(request.GET.get("sayfa"))

    return render(request, "secmelidersler/ortalama_listesi.html", {
        "title": "Dönem Ağırlıklı Ortalamalar",
        "aktif_yil": aktif_yil,
        "sayfa": sayfa,
        "arama": arama,
        "toplam": qs.count(),
    })


@mudur_yardimcisi_required
def ortalama_yukle(request):
    if request.method != "POST":
        return redirect("ortalama_listesi")

    aktif_yil = get_aktif_egitim_yili()
    dosya = request.FILES.get("excel_dosya")
    if not dosya:
        messages.error(request, "Dosya seçilmedi.")
        return redirect("ortalama_listesi")

    from ogrenci.models import Ogrenci

    try:
        satirlar = _excel_satirlari_oku(dosya)
    except Exception as e:
        messages.error(request, f"Dosya okunamadı: {e}")
        return redirect("ortalama_listesi")

    if len(satirlar) < 2:
        messages.warning(request, "Dosya boş veya yalnızca başlık satırı içeriyor.")
        return redirect("ortalama_listesi")

    okulno_idx, a_ort_idx = _sutun_indeksleri_bul(satirlar[0])
    if okulno_idx is None or a_ort_idx is None:
        messages.error(
            request,
            "Başlık satırında 'okulno' ve 'a_ortalama' sütunları bulunamadı. "
            "Lütfen Excel dosyanızın başlık satırını kontrol edin."
        )
        return redirect("ortalama_listesi")

    ogrenci_map = {o.okulno: o for o in Ogrenci.objects.all()}

    olusturulan = guncellenen = bulunamayan = hatali = 0
    bulunamayan_liste = []

    for satir_no, satir in enumerate(satirlar[1:], start=2):
        try:
            ham_okulno = satir[okulno_idx]
            ham_ort    = satir[a_ort_idx]
            if ham_okulno is None or ham_ort is None:
                continue
            okulno   = int(float(str(ham_okulno).strip()))
            a_ort    = round(float(str(ham_ort).strip().replace(",", ".")), 2)
        except (ValueError, TypeError, IndexError):
            hatali += 1
            continue

        ogr = ogrenci_map.get(okulno)
        if ogr is None:
            bulunamayan += 1
            bulunamayan_liste.append(okulno)
            continue

        obj, created = OgrenciOrtalama.objects.update_or_create(
            ogrenci=ogr,
            egitim_yili=aktif_yil,
            defaults={"a_ortalama": a_ort},
        )
        if created:
            olusturulan += 1
        else:
            guncellenen += 1

    ozet = f"{olusturulan} yeni kayıt eklendi, {guncellenen} kayıt güncellendi."
    if bulunamayan:
        ozet += f" {bulunamayan} okul no sistemde bulunamadı."
    if hatali:
        ozet += f" {hatali} satır hatalı format nedeniyle atlandı."

    if olusturulan + guncellenen > 0:
        messages.success(request, ozet)
    else:
        messages.warning(request, ozet)

    if bulunamayan_liste:
        request.session["bulunamayan_okulno"] = bulunamayan_liste[:50]

    return redirect("ortalama_listesi")


@mudur_yardimcisi_required
def ortalama_sil(request, pk):
    obj = OgrenciOrtalama.objects.filter(pk=pk).first()
    if obj:
        obj.delete()
        messages.success(request, f"{obj.ogrenci} kaydı silindi.")
    return redirect("ortalama_listesi")


@mudur_yardimcisi_required
def ortalama_toplu_sil(request):
    if request.method == "POST":
        aktif_yil = get_aktif_egitim_yili()
        sayi, _ = OgrenciOrtalama.objects.filter(egitim_yili=aktif_yil).delete()
        messages.success(request, f"{sayi} kayıt silindi.")
    return redirect("ortalama_listesi")


# ---------------------------------------------------------------------------
# Tasdikname CRUD — Okuma Hakkı Biten Öğrenciler
# ---------------------------------------------------------------------------

@mudur_yardimcisi_required
def tasdikname_listesi(request):
    from ogrenci.models import Ogrenci

    aktif_yil = get_aktif_egitim_yili()
    kayitlar = (
        OgrenciTasdikname.objects
        .select_related("ogrenci")
        .order_by("ogrenci__sinif", "ogrenci__sube", "ogrenci__okulno")
    )

    arama = request.GET.get("q", "").strip()
    arama_sonuclari = []
    if arama:
        from django.db.models import Q
        mevcut_ids = set(kayitlar.values_list("ogrenci_id", flat=True))
        arama_sonuclari = list(
            Ogrenci.objects.filter(
                Q(okulno__icontains=arama)
                | Q(adi__icontains=arama)
                | Q(soyadi__icontains=arama),
                aktif=True,
            ).exclude(pk__in=mevcut_ids).order_by("sinif", "sube", "okulno")[:20]
        )

    return render(request, "secmelidersler/tasdikname_listesi.html", {
        "title": "Tasdikname — Okuma Hakkı Biten Öğrenciler",
        "aktif_yil": aktif_yil,
        "kayitlar": kayitlar,
        "arama": arama,
        "arama_sonuclari": arama_sonuclari,
        "toplam": kayitlar.count(),
    })


@mudur_yardimcisi_required
def tasdikname_ekle(request):
    if request.method != "POST":
        return redirect("tasdikname_listesi")

    from ogrenci.models import Ogrenci
    ogrenci_pk = request.POST.get("ogrenci_pk")
    tarih = request.POST.get("tarih") or None
    aciklama = request.POST.get("aciklama", "").strip()
    aktif_yil = get_aktif_egitim_yili()

    ogr = Ogrenci.objects.filter(pk=ogrenci_pk).first()
    if not ogr:
        messages.error(request, "Öğrenci bulunamadı.")
        return redirect("tasdikname_listesi")

    _, created = OgrenciTasdikname.objects.get_or_create(
        ogrenci=ogr,
        defaults={"egitim_yili": aktif_yil, "tarih": tarih, "aciklama": aciklama},
    )
    if created:
        if ogr.aktif:
            ogr.aktif = False
            ogr.save(update_fields=["aktif"])
        messages.success(
            request,
            f"{ogr.adi} {ogr.soyadi} tasdikname listesine eklendi ve aktif öğrenci "
            "listelerinden çıkarıldı.",
        )
    else:
        messages.warning(request, f"{ogr.adi} {ogr.soyadi} zaten tasdikname listesinde.")

    from django.urls import reverse as _rev
    q = request.POST.get("q", "")
    return redirect(f"{_rev('tasdikname_listesi')}?q={q}")


@mudur_yardimcisi_required
def tasdikname_sil(request, pk):
    if request.method == "POST":
        kayit = OgrenciTasdikname.objects.filter(pk=pk).select_related("ogrenci").first()
        if kayit:
            ogr = kayit.ogrenci
            ad = f"{ogr.adi} {ogr.soyadi}"
            kayit.delete()
            if not ogr.aktif:
                ogr.aktif = True
                ogr.save(update_fields=["aktif"])
            messages.success(
                request, f"{ad} tasdikname listesinden çıkarıldı ve tekrar aktif hale getirildi."
            )
    return redirect("tasdikname_listesi")


# ---------------------------------------------------------------------------
# Sınıf Tekrarı CRUD
# ---------------------------------------------------------------------------

@mudur_yardimcisi_required
def sinif_tekrari_listesi(request):
    from ogrenci.models import Ogrenci

    aktif_yil = get_aktif_egitim_yili()
    kayitlar = (
        OgrenciSinifTekrari.objects
        .select_related("ogrenci")
        .order_by("ogrenci__sinif", "ogrenci__sube", "ogrenci__okulno")
    )
    # Yalnızca AKTİF yıla ait tekrar kayıtları listelenir — aksi hâlde geçmiş bir
    # yılda (örn. 2025-2026) sınıfta kalmış ama o yılı zaten tamamlayıp normal
    # devam eden öğrenciler de bu listede kalıcı olarak görünür (bkz. commit
    # geçmişi).
    if aktif_yil:
        kayitlar = kayitlar.filter(egitim_yili=aktif_yil)

    arama = request.GET.get("q", "").strip()
    arama_sonuclari = []
    if arama:
        from django.db.models import Q
        mevcut_ids = set(kayitlar.values_list("ogrenci_id", flat=True))
        arama_sonuclari = list(
            Ogrenci.objects.filter(
                Q(okulno__icontains=arama)
                | Q(adi__icontains=arama)
                | Q(soyadi__icontains=arama),
                aktif=True,
            ).exclude(pk__in=mevcut_ids).order_by("sinif", "sube", "okulno")[:20]
        )

    return render(request, "secmelidersler/sinif_tekrari_listesi.html", {
        "title": "Sınıf Tekrarı Öğrencileri",
        "aktif_yil": aktif_yil,
        "kayitlar": kayitlar,
        "arama": arama,
        "arama_sonuclari": arama_sonuclari,
        "toplam": kayitlar.count(),
    })


@mudur_yardimcisi_required
def sinif_tekrari_ekle(request):
    if request.method != "POST":
        return redirect("sinif_tekrari_listesi")

    from ogrenci.models import Ogrenci
    ogrenci_pk = request.POST.get("ogrenci_pk")
    aciklama = request.POST.get("aciklama", "").strip()
    aktif_yil = get_aktif_egitim_yili()

    ogr = Ogrenci.objects.filter(pk=ogrenci_pk).first()
    if not ogr:
        messages.error(request, "Öğrenci bulunamadı.")
        return redirect("sinif_tekrari_listesi")

    # `ogrenci` alanı OneToOne olduğundan (bir öğrencinin en fazla bir tekrar kaydı
    # olabilir) `update_or_create` kullanılır — öğrencinin ESKİ bir yıla (örn.
    # 2025-2026) ait kaydı varsa bu, "ekle" işlemiyle AKTİF yıla taşınır/güncellenir;
    # aksi hâlde eski yıl bilgisiyle kalıp öğrenci yanlışlıkla o yılda tekrarcı
    # görünmeye devam eder (bkz. commit geçmişi). Güncellemeden ÖNCEKİ hâli
    # (`mevcut_tekrar`) — ikinci tekrar tespiti için gerekli — burada yakalanır.
    mevcut_tekrar = OgrenciSinifTekrari.objects.filter(ogrenci=ogr).first()

    _, created = OgrenciSinifTekrari.objects.update_or_create(
        ogrenci=ogr,
        defaults={"egitim_yili": aktif_yil, "aciklama": aciklama},
    )

    from secmelidersler.services.sinif_tekrari import ikinci_tekrar_ise_ogrenim_hakkini_sonlandir
    ikinci_mi = ikinci_tekrar_ise_ogrenim_hakkini_sonlandir(ogr, mevcut_tekrar, aktif_yil)

    if ikinci_mi:
        messages.warning(
            request,
            f"{ogr.adi} {ogr.soyadi} — bu İKİNCİ sınıf tekrarı kaydı ({mevcut_tekrar.egitim_yili} → "
            f"{aktif_yil}); öğrenim hakkı tamamlandı kabul edilip tasdikname/ayrılma kaydı otomatik "
            "oluşturuldu, öğrenci pasife alındı.",
        )
    elif created:
        messages.success(request, f"{ogr.adi} {ogr.soyadi} sınıf tekrarı listesine eklendi.")
    else:
        messages.success(request, f"{ogr.adi} {ogr.soyadi} sınıf tekrarı kaydı {aktif_yil} yılına güncellendi.")

    from django.urls import reverse as _rev
    q = request.POST.get("q", "")
    return redirect(f"{_rev('sinif_tekrari_listesi')}?q={q}")


@mudur_yardimcisi_required
def sinif_tekrari_sil(request, pk):
    if request.method == "POST":
        kayit = OgrenciSinifTekrari.objects.filter(pk=pk).select_related("ogrenci").first()
        if kayit:
            ad = f"{kayit.ogrenci.adi} {kayit.ogrenci.soyadi}"
            kayit.delete()
            messages.success(request, f"{ad} sınıf tekrarı listesinden çıkarıldı.")
    return redirect("sinif_tekrari_listesi")


# ─────────────────────────────────────────────
# Önceki Dönemden Getir
# ─────────────────────────────────────────────


def _onceki_yil_kontrol(request):
    aktif_yil = get_aktif_egitim_yili()
    if not aktif_yil:
        messages.error(request, "Aktif eğitim-öğretim yılı tanımlı değil.")
        return None, None
    onceki_yil = donem_kopyala.onceki_egitim_yili(aktif_yil)
    if not onceki_yil:
        messages.error(request, "Önceki bir eğitim-öğretim yılı bulunamadı.")
        return aktif_yil, None
    return aktif_yil, onceki_yil


@mudur_yardimcisi_required
@require_POST
def sinif_toplam_saat_onceki_getir(request):
    aktif_yil, onceki_yil = _onceki_yil_kontrol(request)
    if onceki_yil:
        n = donem_kopyala.sinif_toplam_saat_kopyala(onceki_yil, aktif_yil)
        messages.success(request, f"{onceki_yil} yılından {n} sınıf seviyesi toplam saati getirildi.")
    return redirect("secmeli_sinif_toplam_saat")


@mudur_yardimcisi_required
@require_POST
def ortak_ders_onceki_getir(request):
    aktif_yil, onceki_yil = _onceki_yil_kontrol(request)
    if onceki_yil:
        eklenen, atlanan = donem_kopyala.ortak_ders_kopyala(onceki_yil, aktif_yil)
        msg = f"{onceki_yil} yılından {eklenen} zorunlu ders getirildi."
        if atlanan:
            msg += f" ({atlanan} tanesi zaten vardı, atlandı.)"
        messages.success(request, msg)
    return redirect("secmeli_ortak_ders_listesi")


@mudur_yardimcisi_required
@require_POST
def secmeli_grup_onceki_getir(request):
    aktif_yil, onceki_yil = _onceki_yil_kontrol(request)
    if onceki_yil:
        eklenen_grup, eklenen_ders, atlanan = donem_kopyala.secmeli_grup_kopyala(onceki_yil, aktif_yil)
        msg = f"{onceki_yil} yılından {eklenen_grup} grup, {eklenen_ders} ders getirildi."
        if atlanan:
            msg += f" ({atlanan} grup zaten vardı, atlandı.)"
        messages.success(request, msg)
    return redirect("secmeli_grup_listesi")


@mudur_yardimcisi_required
@require_POST
def alan_onceki_getir(request):
    aktif_yil, onceki_yil = _onceki_yil_kontrol(request)
    if onceki_yil:
        eklenen_alan, eklenen_ders, atlanan, eslesmeyen = donem_kopyala.alan_kopyala(onceki_yil, aktif_yil)
        msg = f"{onceki_yil} yılından {eklenen_alan} alan, {eklenen_ders} alan dersi getirildi."
        if atlanan:
            msg += f" ({atlanan} alan zaten vardı, atlandı.)"
        if eslesmeyen:
            msg += (
                f" {eslesmeyen} ders eşleşmedi — önce 'Seçmeli Ders Grupları' ekranından "
                "önceki dönemi getirin."
            )
        messages.success(request, msg)
    return redirect("secmeli_alan_listesi")


